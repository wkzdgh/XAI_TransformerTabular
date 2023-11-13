import torch
import time
import os
import torch.optim as optim
import pandas as pd
from torch import nn
import random
import torchmetrics
import numpy as np
import csv
from sklearn.preprocessing import MinMaxScaler
import models
import shap
from openpyxl import Workbook, load_workbook

def create_xlsx(path, name_dataset):
    if not os.path.exists(path):
        workbook = Workbook()
        workbook.active.title = name_dataset
        workbook.save(path)

def create_sheet(path, name_dataset):
    wb = load_workbook(path)
    if name_dataset not in wb.sheetnames:
        wb.create_sheet(name_dataset)
        wb.save(path)

def export_explanation_to_excel(result_path, name_dataset, dict_explanation):
    path = result_path + os.sep + "explanation.xlsx"
    create_xlsx(path, name_dataset)
    create_sheet(path, name_dataset)

    writer = pd.ExcelWriter(path=path, mode="a", engine="openpyxl", if_sheet_exists="overlay")
    
    row = 1
    for key in dict_explanation:
        column = 1
        writer.sheets[name_dataset].cell(row=row, column=column).value = key
        row += 1
        for feature, expl in dict_explanation[key]:
            writer.sheets[name_dataset].cell(row=row, column=column).value = feature
            row += 1
            expl["MeanRelevance"] = expl.mean(axis=1)
            expl.to_excel(writer, sheet_name=name_dataset, startrow=row-1, startcol=column-1)
            row -= 1
            column += expl.shape[1] + 2 #sumar número de columnas más 2 de espacio entre tablas
            num_rows = expl.shape[0] + 4
        row += num_rows #sumar número de filas más 4 de espacio entre tablas
    writer.close()
    return 0

def export_accuracy_to_excel(result_path, name_dataset, dict_accuracy, nfeats, name_folders, name_nfeats):
    path = result_path + os.sep + "accuracy.xlsx"
    create_xlsx(path, name_dataset)
    create_sheet(path, name_dataset)
    
    writer = pd.ExcelWriter(path=path, mode="a", engine="openpyxl", if_sheet_exists="overlay")

    dataframes = []
    df_medias = pd.DataFrame(columns=name_nfeats, index=dict_accuracy.keys())
    for n in range(0, nfeats):
        table_nfeat = np.zeros(shape=(len(dict_accuracy), len(name_folders)))
        for i, key_model in enumerate(dict_accuracy):
            table_nfeat[i] = dict_accuracy[key_model].iloc[n]
        df = pd.DataFrame(table_nfeat, columns=name_folders, index=dict_accuracy.keys())
        df["MeanAccuracy"] = df.mean(axis=1)
        df_medias[name_nfeats[n]] = df["MeanAccuracy"]       #df_medias.loc[:, name_nfeats[n]] = df["MeanAccuracy"]
        dataframes.append(df)
    dataframes.append(df_medias)

    column = 1
    row = 1
    for i, name_df in enumerate(name_nfeats):            
        writer.sheets[name_dataset].cell(row=row, column=column).value = name_df
        row += 1
        dataframes[i].to_excel(writer, sheet_name=name_dataset, startrow=row-1, startcol=column-1)
        row += df.shape[0] + 4    

    writer.sheets[name_dataset].cell(row=row, column=column).value = "MeanAccuracy"
    row += 1
    dataframes[i+1].to_excel(writer, sheet_name=name_dataset, startrow=row-1, startcol=column-1)

    writer.close()
    return 0

def predict_explain_models(model, X_train, X_test, y_test, device, rf):
    y_pred = torch.from_numpy(model.predict(X_test)).to(device)
    y_pred = torch.unsqueeze(y_pred, dim=1)

    if rf == True:
        explainer = shap.TreeExplainer(model)
        shap_values = explainer(X_test)
        expls = [np.transpose(np.transpose(shap_values.values)[y_pred[i]])[i] for i in range(0, y_pred.shape[0])]
    else:
        explainer = shap.KernelExplainer(model.predict_proba, shap.kmeans(X_train, 10))
        shap_values = explainer.shap_values(X_train) #(X_test) #TODO: la métrica de accuracy se saca con X_test pero la explicación se saca con X_train
        expls = [shap_values[y_pred[i]][i] for i in range(0, y_pred.shape[0])]
    expls = torch.from_numpy(np.array(expls)).to(device)
    metric_value = model.score(X_test, y_test)
    print("\nMean Accuracy in test data: ", str(metric_value))

    return expls, metric_value

def join_cat_cont(trainloader, testloader):
    if trainloader.dataset.cat.shape[1] > 0:
        if trainloader.dataset.cont.shape[1] > 0:
            X_train = np.concatenate([trainloader.dataset.cat, trainloader.dataset.cont], axis=1)
            X_test = np.concatenate([testloader.dataset.cat, testloader.dataset.cont], axis=1)
        else:
            X_train = trainloader.dataset.cat
            X_test = testloader.dataset.cat
    elif trainloader.dataset.cont.shape[1] > 0:
        X_train = trainloader.dataset.cont
        X_test = testloader.dataset.cont
    else:
        print("ERROR - ...")

    y_train = trainloader.dataset.y 
    y_test = testloader.dataset.y

    return X_train, y_train, X_test, y_test

def delete_feature(trainloader, testloader, feature_deleted):
    if feature_deleted < trainloader.dataset.cat.shape[1]:
        trainloader.dataset.cat = np.delete(trainloader.dataset.cat, feature_deleted, axis=1)
        trainloader.dataset.cat_mask = np.delete(trainloader.dataset.cat_mask, feature_deleted, axis=1)
        i = trainloader.dataset.dataCat.pop(feature_deleted)
        trainloader.dataset.attribute_names.pop(trainloader.dataset.attribute_names.index(i[1]))
        testloader.dataset.cat = np.delete(testloader.dataset.cat, feature_deleted, axis=1)
        testloader.dataset.cat_mask = np.delete(testloader.dataset.cat_mask, feature_deleted, axis=1)
        i = testloader.dataset.dataCat.pop(feature_deleted)
        testloader.dataset.attribute_names.pop(testloader.dataset.attribute_names.index(i[1]))    
    else:
        trainloader.dataset.cont = np.delete(trainloader.dataset.cont, feature_deleted-trainloader.dataset.cat.shape[1], axis=1)
        trainloader.dataset.cont_mask = np.delete(trainloader.dataset.cont_mask, feature_deleted-trainloader.dataset.cat.shape[1], axis=1)
        i = trainloader.dataset.dataCont.pop(feature_deleted-trainloader.dataset.cat.shape[1])
        trainloader.dataset.attribute_names.pop(trainloader.dataset.attribute_names.index(i[1]))
        testloader.dataset.cont = np.delete(testloader.dataset.cont, feature_deleted-testloader.dataset.cat.shape[1], axis=1)
        testloader.dataset.cont_mask = np.delete(testloader.dataset.cont_mask, feature_deleted-testloader.dataset.cat.shape[1], axis=1)
        i = testloader.dataset.dataCont.pop(feature_deleted-testloader.dataset.cat.shape[1])
        testloader.dataset.attribute_names.pop(testloader.dataset.attribute_names.index(i[1]))
    
    return trainloader, testloader, trainloader.dataset.cat.shape[1] + trainloader.dataset.cont.shape[1]  


def cross_validation_process(trainloader, testloader, y_dim, opt, device, criterion):
    cat_dims = [trainloader.dataset.dataCat[i][2] for i in range(len(trainloader.dataset.dataCat))] 
    cat_dims = np.append(np.array([1]),np.array(cat_dims)).astype(int) #Appending 1 for CLS token, this is later used to generate embeddings.
    nfeat = trainloader.dataset.cat.shape[1] + trainloader.dataset.cont.shape[1]  

    model = models.SAINT(
        num_features = nfeat + 1,
        categories = tuple(cat_dims), 
        num_continuous = trainloader.dataset.cont.shape[1],       
        dim = opt.embedding_size,                           
        dim_out = 1,                       
        depth = opt.transformer_depth,                       
        heads = opt.attention_heads,                         
        attn_dropout = opt.attention_dropout,             
        ff_dropout = opt.ff_dropout,                  
        mlp_hidden_mults = (4, 2),       
        cont_embeddings = opt.cont_embeddings,
        attentiontype = opt.attentiontype,
        final_mlp_style = opt.final_mlp_style,
        y_dim = y_dim
    )
    
    model.to(device)

    optimizer, scheduler = select_optimizer(model, opt.optimizer, opt.scheduler, opt.epochs, opt.lr)

    
    #model = train(model, trainloader, opt.task, 3, device, criterion, opt.optimizer, optimizer, scheduler)
    model = train(model, trainloader, opt.task, opt.epochs, device, criterion, opt.optimizer, optimizer, scheduler)
    print("\tModelo entrenado, calculando métricas...")
    
    explainator = models.ExplainationGenerator(model)
    expls, y_pred, y_gts = predict_all(model, explainator, testloader, device)
    metric, metric_name = create_metric(opt.task, y_dim, device)
    metric_value = metric(torch.squeeze(y_pred), torch.squeeze(y_gts))
    print("\t" + metric_name + ": " + str(torch.Tensor.numpy(metric_value.cpu())))

    #TODO: la métrica de accuracy se saca con testloader pero la explicación se saca con trainloader
    expls, y_pred, y_gts = predict_all(model, explainator, trainloader, device)
    return expls, metric_value


"""def drop_row_column(tensor, index, row_colum):
    idx_get = [i for i in range(0, tensor.shape[row_colum])]
    if row_colum == 0:
        tensor = tensor[idx_get != index, :]
    elif row_colum == 1:
        tensor = tensor[:, (idx_get != 1)]
    return tensor"""


def embed_data_mask(x_categ, x_cont, cat_mask, con_mask,model):#,vision_dset=False):
    device = x_cont.device
    x_categ = x_categ + model.categories_offset.type_as(x_categ)
    x_categ_enc = model.embeds(x_categ)
    n1,n2 = x_cont.shape
    _, n3 = x_categ.shape
    if model.cont_embeddings == 'MLP':
        x_cont_enc = torch.empty(n1,n2, model.dim)
        for i in range(model.num_continuous):
            x_cont_enc[:,i,:] = model.simple_MLP[i](x_cont[:,i])
    else:
        raise Exception('This case should not work!')    

    x_cont_enc = x_cont_enc.to(device)
    cat_mask_temp = cat_mask + model.cat_mask_offset.type_as(cat_mask)
    con_mask_temp = con_mask + model.con_mask_offset.type_as(con_mask)

    cat_mask_temp = model.mask_embeds_cat(cat_mask_temp)
    con_mask_temp = model.mask_embeds_cont(con_mask_temp)
    x_categ_enc[cat_mask == 0] = cat_mask_temp[cat_mask == 0]
    x_cont_enc[con_mask == 0] = con_mask_temp[con_mask == 0]

    return x_categ, x_categ_enc, x_cont_enc


def select_criterion(y_dim, task, device):
    if y_dim == 2 and task == 'binary':
        criterion = nn.CrossEntropyLoss().to(device)
    elif y_dim > 2 and task == 'multiclass':
        criterion = nn.CrossEntropyLoss().to(device)
    elif task == 'regression':
        criterion = nn.MSELoss().to(device)
    else:
        raise'case not written yet' #TODO
    return criterion


def get_scheduler(scheduler_type, epochs, optimizer):
    if scheduler_type == 'cosine':
        scheduler = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer, epochs)
    elif scheduler_type == 'linear':
        scheduler = torch.optim.lr_scheduler.MultiStepLR(optimizer,
                                      milestones=[epochs // 2.667, epochs // 1.6, epochs // 1.142], gamma=0.1)
    return scheduler


def select_optimizer(model, optimizer_type, scheduler_type, epochs, lr):
    if optimizer_type == 'SGD':
        optimizer = optim.SGD(model.parameters(), lr=lr,
                            momentum=0.9, weight_decay=5e-4)
        scheduler = get_scheduler(scheduler_type, epochs, optimizer)
    elif optimizer_type == 'Adam':
        optimizer = optim.Adam(model.parameters(),lr=lr)
    elif optimizer_type == 'AdamW':
        optimizer = optim.AdamW(model.parameters(),lr=lr)

    if optimizer_type == 'SGD':
        return optimizer, scheduler
    else:
        return optimizer, None    


def create_metric(task, num_classes, device):
    if task in ['binary','multiclass']:
        metric = torchmetrics.Accuracy(task=task, num_classes=num_classes).to(device)
        metric_name = "Accuracy"
    else:
        metric = torchmetrics.MeanSquaredError(squared=False, num_outputs=1)
        metric_name = "RMSE"
    return metric, metric_name


def train(model, trainloader, task, epochs, device, criterion, optimizer_type, optimizer, scheduler):
    for epoch in range(epochs):
        if epoch % 20 == 0: print("\tEpoch " + str(epoch) + ": ")
        model.train()
        running_loss = 0.0
        for i, data in enumerate(trainloader, 0):
            optimizer.zero_grad()

            x_categ, x_cont, y_gts, cat_mask, con_mask = data[0].to(device), data[1].to(device),data[2].to(device),data[3].to(device),data[4].to(device)
            _ , x_categ_enc, x_cont_enc = embed_data_mask(x_categ, x_cont, cat_mask, con_mask, model)      
            reps = model.transformer(x_categ_enc, x_cont_enc)
            y_reps = reps[:,0,:]
            y_outs = model.mlpfory(y_reps)

            if task == 'regression':
                loss = criterion(y_outs,y_gts) 
            else:
                loss = criterion(y_outs,y_gts.squeeze()) 
            loss.backward()
            optimizer.step()
            if optimizer_type == 'SGD':
                scheduler.step()
            running_loss += loss.item()
    return model


def predict_one_data(model, task, num_classes, explainator, testloader, device):
    metric, metric_name = create_metric(task, num_classes, device)
    batch_id = random.randint(0, len(testloader) - 1)
    model.eval()
    for i, data in enumerate(testloader, 0):
        if i == batch_id:
            data_id = random.randint(0, data[0].shape[0]-1)

            x_categ, x_cont, y_gts, cat_mask, con_mask = data[0].to(device), data[1].to(device),data[2].to(device),data[3].to(device),data[4].to(device)
            _ , x_categ_enc, x_cont_enc = embed_data_mask(x_categ, x_cont, cat_mask, con_mask, model)      
            reps = model.transformer(x_categ_enc, x_cont_enc)
            y_reps = reps[:,0,:]
            y_outs = model.mlpfory(y_reps)
            y_pred = torch.from_numpy(np.argmax(y_outs.cpu().data.numpy(), axis=1)).to(device)
            metric_value = metric(y_pred, torch.squeeze(y_gts))
            print(str(metric.compute()))
            print('TEST %s: %.3f      -      %s' % (metric_name, metric_value, time.asctime(time.localtime(time.time())) ))

            print("\nCalculando explicación para el dato " + str(data_id) + " del batch " + str(batch_id))
            x_categ, x_cont, y_gts, cat_mask, con_mask = data[0].to(device), data[1].to(device),data[2].to(device),data[3].to(device),data[4].to(device)
            _ , x_categ_enc, x_cont_enc = embed_data_mask(x_categ, x_cont, cat_mask, con_mask,model)
            x_categ_enc_item = torch.unsqueeze(x_categ_enc[data_id], dim=0) 
            x_cont_enc_item = torch.unsqueeze(x_cont_enc[data_id], dim=0) 
            if explainator == None:
                reps = model.transformer(x_categ_enc_item , x_cont_enc_item)
                y_reps = reps[:,0,:]
                output = model.mlpfory(y_reps) 
                print("ERROR - No se ha pasado un explicador para explicar el modelo")
            else:
                expl, output = explainator.generateExplanation(x_categ_enc_item , x_cont_enc_item, device)    
                expl = (expl - expl.min()) / (expl.max() - expl.min())
                df_expl = pd.DataFrame(data=expl.cpu().detach().numpy(), columns=testloader.dataset.attribute_names).astype("float")
            
            pDoutput = pd.DataFrame(output.cpu().detach().numpy())
            explanation = df_expl.transpose()
            y_true = (int(torch.unsqueeze(y_gts[data_id], dim=0).cpu().numpy()))
            y_pred = output.argmax(dim=-1).item()

            return pDoutput, explanation, y_true, y_pred, x_categ[data_id], x_cont[data_id]


def get_idx_elements_classes(x):
    dic = {}
    for i in range(0, x.shape[0]):
        if x[i].item() not in dic:
            dic[x[i].item()] = [i]
        else:
            dic[x[i].item()].append(i)
    return dic


def get_relevances_from_idx(x, expls):
    relevance_classes = {}
    for i, classes in enumerate(x):
        for j, id in enumerate(x[classes]):
            if classes not in relevance_classes:
                relevance_classes[classes] = [expls[id]]
            else:
                relevance_classes[classes].append(expls[id])
    return relevance_classes


#TODO -> REVISAR: CAMBIAR LA PRIMERA PARTE PARA QUE SE USE LA FUNCIÓN DROP_ROW_COLUMN EN LUGAR DE PASAR
#DE TENSOR A NUMPY Y DE NUMPY A TENSOR
def get_metrics_explanation(expls, y_pred, y_gts, task, num_classes, device):
    metric, metric_name = create_metric(task, num_classes, device)
    metric_value = metric(torch.squeeze(y_pred), torch.squeeze(y_gts))

    expls_correct = expls.cpu().detach().numpy()
    y_correct_list = []
    i_deleted = []
    for i in range(0, y_gts.shape[0]):
        if y_gts[i].item() != y_pred[i].item():
            i_deleted.append(i)
        else:
            y_correct_list.append(y_gts[i].item())
    expls_correct = np.delete(expls_correct, i_deleted, 0)
    y_correct = torch.from_numpy(np.array(y_correct_list))
    y_correct = torch.unsqueeze(y_correct, dim=1)
    expls_correct = torch.from_numpy(expls_correct)

    idx_y_gts = get_idx_elements_classes(y_gts)
    idx_y_pred = get_idx_elements_classes(y_pred)
    idx_correct = get_idx_elements_classes(y_correct)
    
    relevance_classes_y_gts = get_relevances_from_idx(idx_y_gts, expls)
    relevance_classes_y_pred = get_relevances_from_idx(idx_y_pred, expls)
    relevance_classes_y_correct = get_relevances_from_idx(idx_correct, expls_correct)

    mean_feature_relevance = expls.mean(dim=0)
    mean_feature_relevance_correct = expls_correct.mean(dim=0)
    mean_relevance_classes_y_gts = [torch.stack(relevance_classes_y_gts[elem]).mean(dim=0).cpu().detach().numpy().tolist() for elem in relevance_classes_y_gts]
    mean_relevance_classes_y_pred = [torch.stack(relevance_classes_y_pred[elem]).mean(dim=0).cpu().detach().numpy().tolist() for elem in relevance_classes_y_pred]
    mean_relevance_classes_y_correct = [torch.stack(relevance_classes_y_correct[elem]).mean(dim=0).cpu().detach().numpy().tolist() for elem in relevance_classes_y_correct]

    return metric_name, metric_value, mean_feature_relevance, mean_feature_relevance_correct, mean_relevance_classes_y_gts, mean_relevance_classes_y_pred, mean_relevance_classes_y_correct


def write_metrics_file(file_name, mode, model_name, num_examples, cat_names, cont_names, metric_name, metric_value, mean_feature_relevance, mean_feature_relevance_correct, mean_relevance_classes_y_gts, mean_relevance_classes_y_pred, mean_relevance_classes_y_correct):
    file = open(file_name + os.sep +  "metrics.txt", mode)
    wr = csv.writer(file, delimiter="\t")
    mms = MinMaxScaler()

    if mode == "w":
        file.write("---------------------------------------------------------------------\n\n")
    else:
        file.write("\n---------------------------------------------------------------------\n\n")

    file.write("Número de muestras del dataset de test: ")
    file.write(str(num_examples))
    file.write("\n\n")

    file.write("Categorical features: ")
    file.write(" ".join(cat_names))
    file.write("\n\n")

    file.write("Continuous features: ")
    file.write(" ".join(cont_names))
    file.write("\n\n")

    file.write("Model: ")
    file.write(model_name)
    file.write("\n\n")

    file.write(metric_name)
    file.write(": ")
    file.write(str(torch.Tensor.numpy(metric_value.cpu())))
    file.write("\n\n")

    file.write("Mean feature relevance with all examples and all classes: ")
    file.write(str(mms.fit_transform(mean_feature_relevance.cpu().detach().numpy().reshape(-1, 1)).reshape(1, -1))) #el reshape(-1, 1) es porque es solo una línea
    file.write("\n\n")

    file.write("Mean feature relevance only with well predicted examples and all classes: ")
    file.write(str(mms.fit_transform(mean_feature_relevance_correct.cpu().detach().numpy().reshape(-1, 1)).reshape(1, -1))) #el reshape(-1, 1) es porque es solo una línea
    file.write("\n\n")

    file.write("Mean feature relevance per classes (according to the correct classes): \n")
    _ = [mms.partial_fit(np.array(fila).reshape(-1, 1)) for fila in mean_relevance_classes_y_gts]
    scaled = [mms.transform(np.array(fila).reshape(-1, 1)).reshape(1, -1) for fila in mean_relevance_classes_y_gts]
    wr.writerows(scaled)
    file.write("\n")

    file.write("Mean feature relevance per classes (according to the predicted classes): \n")
    _ = [mms.partial_fit(np.array(fila).reshape(-1, 1)) for fila in mean_relevance_classes_y_pred]
    scaled = [mms.transform(np.array(fila).reshape(-1, 1)).reshape(1, -1) for fila in mean_relevance_classes_y_pred]
    wr.writerows(scaled)
    file.write("\n")

    file.write("Mean feature relevance per classes (only with well predicted examples): \n")
    _ = [mms.partial_fit(np.array(fila).reshape(-1, 1)) for fila in mean_relevance_classes_y_correct]
    scaled = [mms.transform(np.array(fila).reshape(-1, 1)).reshape(1, -1) for fila in mean_relevance_classes_y_correct]
    wr.writerows(scaled)

    file.close()


def predict_all(model, explainator, dataset, device):
    model.eval()
    for i, data in enumerate(dataset, 0):
        x_categ, x_cont, y_gts, cat_mask, con_mask = data[0].to(device), data[1].to(device),data[2].to(device),data[3].to(device),data[4].to(device)
        _ , x_categ_enc, x_cont_enc = embed_data_mask(x_categ, x_cont, cat_mask, con_mask, model)
        
        if explainator == None:
            reps = model.transformer(x_categ_enc , x_cont_enc)
            y_reps = reps[:,0,:]
            output = model.mlpfory(y_reps) 
            print("ERROR - No se ha pasado un explicador para explicar el modelo")
        else:
            expls, outputs = explainator.generateExplanation_all(x_categ_enc, x_cont_enc, device)
            expls = expls[:, 1:]
            y_pred = torch.from_numpy(np.argmax(outputs.cpu().data.numpy(), axis=1)).to(device)
            y_pred = torch.unsqueeze(y_pred, dim=1)
            return expls, y_pred, y_gts

     
def safe_divide(a, b):
    den = b.clamp(min=1e-9) + b.clamp(max=1e-9)
    den = den + den.eq(0).type(den.type()) * 1e-9
    return a / den * b.ne(0).type(b.type())


def forward_hook(self, input, output):
    if type(input[0]) in (list, tuple):
        self.X = []
        for i in input[0]:
            x = i.detach()
            x.requires_grad = True
            self.X.append(x)
    else:
        self.X = input[0].detach()
        self.X.requires_grad = True

    self.Y = output